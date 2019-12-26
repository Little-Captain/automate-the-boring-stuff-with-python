#!/usr/bin/env python

# 向会员发送会费提醒电子邮件
# 在较高的层面上，程序要做的事:
# 1 从 Excel 电子表格中读取数据
# 2 找出上个月没有交费的所有会员
# 3 找到他们的电子邮件地址，向他们发送针对个人的提醒
# 代码需要做到以下几点:
# 1 用 openpyxl 模块打开并读取 Excel 文档的单元格（处理 Excel 文件）
# 2 创建一个字典，包含会费超期的会员
# 3 调用 smtplib.SMTP()、ehlo()、starttls() 和 login()，登录 SMTP 服务器
# 4 针对会费超期的所有会员，调用 sendmail() 方法，发送针对个人的电子邮件提醒

# 第 1 步: 打开 Excel 文件
# 第 2 步: 查找所有未付成员
# 第 3 步: 发送定制的电子邮件提醒

import openpyxl
import smtplib
import sys
import os
import sqlite3

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']

lastCol = sheet.max_column
lastestMonth = sheet.cell(row=1, column=lastCol).value

# Check each member's payment status.
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Log in to email account.
smtpObj = smtplib.SMTP('smtp.qq.com', 587)
smtpObj.ehlo()
smtpObj.starttls()

username = '1203118412@qq.com'
password = None
# 连接数据库获取邮箱登录密码
connect = sqlite3.connect(r'/mnt/c/Users/Little-Captain/.sqlite/lc.db')
cursor = connect.cursor()

cursor.execute(
    "select password from user where type = 'smtp-imap' and username = %r" % username)
values = cursor.fetchall()
if len(values) == 1 and len(values[0]) == 1:
    password = os.popen('encrypt.exe -d %r' % values[0][0]).read()

if password == None:
    sys.exit(-1)
else:
    password = password.strip()

smtpObj.login(username, password)

# Send out reminder emails.
for (name, email) in unpaidMembers.items():
    body = "Subject: %s dues unpaid.\nDear %s, \nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!'" % (lastestMonth, name, lastestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail(username, email, body)

    if sendmailStatus != []:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))

smtpObj.quit()

wb.close()
