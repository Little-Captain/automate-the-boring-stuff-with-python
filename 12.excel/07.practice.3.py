#!/usr/bin/env python

# 电子表格单元格翻转程序
# 编写一个程序，翻转电子表格中行和列的单元格

import openpyxl

wb = openpyxl.load_workbook('example.xlsx', read_only=True)
sheet = wb.active

wb_new = openpyxl.Workbook()
sheet_new = wb_new.active
sheet_new.title = 'reversed'

for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        sheet_new.cell(j, i, sheet.cell(i, j).value)

wb_new.save('example_r.xlsx')
wb.close()
wb_new.close()
