#!/usr/bin/env python

# 电子表格到文本文件
# 编写一个程序，执行前一个程序相反的任务。该程序应该打开一个电子表格，
# 将列 A 中的单元格写入一个文本文件，将列 B 中的单元格写入另一个文本文件，以此类推。

import openpyxl
from openpyxl.utils import get_column_letter
import pprint

wb = openpyxl.load_workbook('files.xlsx')
sheet = wb.active

for column in range(1, sheet.max_column + 1):
    lines = []
    for row in range(1, sheet.max_row + 1):
        lines.append(sheet.cell(row, column).value)
    lines = list(filter(lambda x: x != None, lines))
    f = open('%dr.txt' % column, 'w')
    f.writelines(lines)
    f.close()
    lines.clear()

wb.close()

