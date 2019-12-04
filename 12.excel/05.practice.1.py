#!/usr/bin/env python

# 乘法表
# 创建程序 multiplicationTable.py，从命令行接受数字 N，在一个 Excel 电子表格中创建一个 N×N 的乘法表
# 行1 和列A 应该用做标签，应该使用粗体
# py multiplicationTable.py 6

import sys
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# 1. 获取 N
N = int(sys.argv[1])

# 2. 创建表格
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet.title = '{} x {}'.format(N, N)

# 3. 写入数据和设置格式
boldFont = Font(bold=True)
# sheet['A1'].font = boldFont
for i in range(1, N + 2):  # 行
    for j in range(1, N + 2):  # 列
        coordinate = get_column_letter(j) + str(i)
        if i == 1 and j != 1:
            sheet[coordinate].font = boldFont
            sheet[coordinate] = j - 1
        elif i != 1 and j == 1:
            sheet[coordinate].font = boldFont
            sheet[coordinate] = i - 1
        elif i == 1 and j == 1:
            pass
        else:
            sheet[coordinate] = (i - 1) * (j - 1)

sheet.freeze_panes = 'B2'

# 4. 保存表格
wb.save('multi.xlsx')
