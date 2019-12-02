#!/usr/bin/env python

# 更新一个电子表格

# 程序做下面的事情:
# 1 循环遍历所有行
# 2 如果该行是 Garlic、Celery 或 Lemons，更新价格

# 代码需要做下面的事情:
# 1 打开电子表格文件
# 2 针对每一行，检查列 A 的值是不是 Celery、Garlic 或 Lemon
# 3 如果是，更新列 B 中的价格
# 4 将该电子表格保存为一个新文件（这样就不会丢失原来的电子表格，以防万一）

# Celery     1.19
# Garlic     3.07
# Lemon      1.27

# 第 1 步: 利用更新信息建立数据结构
# 第 2 步: 检查所有行，更新不正确的价格

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# The product types and their updated prices
PRICES_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

# Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row): # skip the first row
    productName = sheet.cell(row=rowNum, column=1).value
    if productName in PRICES_UPDATES:
        print('updated {} {}'.format(rowNum, PRICES_UPDATES[productName]))
        sheet.cell(row=rowNum, column=2).value = PRICES_UPDATES[productName]

wb.save('updatedProduceSales.xlsx')
