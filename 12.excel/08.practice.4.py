#!/usr/bin/env python

# 文本文件到电子表格
# 编写一个程序，读入几个文本文件的内容（可以自己创造这些文本文件），
# 并将这些内容插入一个电子表格，每行写入一行文本。第一个文本文件中的行
# 将写入列 A 中的单元格，第二个文本文件中的行将写入列 B 中的单元格，以此类推

import openpyxl
from openpyxl.utils import get_column_letter
import sys

filenames = sys.argv[1:]

wb = openpyxl.Workbook()
sheet = wb.active

for i in range(len(filenames)):
    filename = filenames[i]
    column = get_column_letter(i + 1)
    with open(filename, 'r') as f:
        lines = f.readlines()
        for j in range(len(lines)):
            coordinate = column + str(j + 1)
            sheet[coordinate] = lines[j]

wb.save('files.xlsx')
wb.close()
