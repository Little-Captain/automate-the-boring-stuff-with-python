#!/usr/bin/env python

# 设置单元格的字体风格
# 设置某些单元格行或列的字体风格，可以帮助你强调电子表格中重点的区域
# 为了定义单元格的字体风格，需要从 openpyxl.styles 模块导入 Font() 和 Style() 函数
import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb['Sheet']
italic24Font = Font(size=24, italic=True)
sheet['A1'].font = italic24Font
sheet['A1'] = 'Hello World!'
wb.save('styled.xlsx')

print('--------------------')

# Font 对象
# Font 对象的 style 属性影响文本在单元格中的显示方式
# 要设置字体风格属性，就向 Font() 函数传入关键字参数
# Font        style      属性的关键字参数
# name        字符串      字体名称，诸如'Calibri' 或 'Times New Roman'
# size        整型        字体大小点数
# bold        布尔型      True 表示粗体
# italic      布尔型      True 表示斜体

# 公式
# 公式以一个等号开始，可以配置单元格，让它包含通过其他单元格计算得到的值
# 利用 openpyxl 模块，用编程的方式在单元格中添加公式，就像添加普通的值一样

wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save('writeFormula.xlsx')

print('--------------------')

# 如果你希望看到该公式的计算结果，而不是原来的公式，就必须将 load_workbook()
# 的 data_only 关键字参数设置为 True。这意味着 Workbook 对象要么显示公式，
# 要么显示公式的结果，不能兼得(但是针对一个电子表格文件，可以加载多个 Workbook 对象)


wbFormulas = openpyxl.load_workbook('writeFormula.xlsx')
sheet = wbFormulas.active
print(sheet['A3'].value)

print('--------------------')

wbDataOnly = openpyxl.load_workbook('writeFormula.xlsx', data_only=True)
sheet = wbDataOnly.active
print(sheet['A3'].value)  # 和上面的结果一样，不知道为什么

# Excel 公式为电子表格提供了一定程度的编程能力，但对于复杂的任务，很快就会失去控制
# Python 代码的可读性比 Excel 公式更好

# 调整行和列

# 设置行高和列宽
# Worksheet 对象有 row_dimensions 和 column_dimensions 属性，控制行高和列宽

wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')
# 行的高度可以设置为 0 到 409 之间的整数或浮点值。这个值表示高度的点数。
# 一点等于 1/72 英寸。默认的行高是 12.75。列宽可以设置为 0 到 255 之间
# 的整数或浮点数。这个值表示使用默认字体大小时(11 点)，单元格可以显示的字符数
# 默认的列宽是 8.43 个字符。列宽为零或行高为零，将使单元格隐藏。

# 合并和拆分单元格
# 利用 merge_cells() 工作表方法，可以将一个矩形区域中的单元格合并为一个单元格
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')
# merge_cells() 的参数是一个字符串，表示要合并的矩形区域左上角和右下角的单元格
# 要设置这些合并后单元格的值，只要设置这一组合并单元格左上角的单元格的值

# 要拆分单元格，就调用 unmerge_cells() 工作表方法
wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active
sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')
wb.save('ummerged.xlsx')

# 冻结窗格
# 对于太大而不能一屏显示的电子表格，“冻结”顶部的几行或最左边的几列，是很有帮助的
# 冻结的列或行表头，就算用户滚动电子表格，也是始终可见的。这称为“冻结窗格”
# 在 OpenPyXL 中，每个 Worksheet 对象都有一个 freeze_panes 属性，
# 可以设置为一个 Cell 对象或一个单元格坐标的字符串
# 请注意，单元格上边的所有行和左边的所有列都会冻结，但单元格所在的行和列不会冻结。

# 冻结窗格的例子
# freeze_panes 的设置             冻结的行和列
# sheet.freeze_panes = 'A2'      行 1
# sheet.freeze_panes = 'B1'      列 A
# sheet.freeze_panes = 'C1'      列 A 和列 B
# sheet.freeze_panes = 'C2'      行 1 和列 A 和列 B
# sheet.freeze_panes = 'A1' 或   没有冻结窗格
# sheet.freeze_panes = None

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2'
wb.save('freezeExample.xlsx')

# 图表
# openpyxl 支持利用工作表中单元格的数据，创建条形图、折线图、散点图和饼图
# 要创建图表，需要做下列事情:
# 1．从一个矩形区域选择的单元格，创建一个 Reference 对象
# 2．通过传入 Reference 对象，创建一个 Series 对象
# 3．创建一个 Chart 对象
# 4．将 Series 对象添加到 Chart 对象
# 5．可选地设置 Chart 对象的 drawing.top、drawing.left、drawing.width
#    和 drawing.height 变量
# 6．将 Chart 对象添加到 Worksheet 对象
# Reference 对象需要一些解释
# Reference 对象是通过调用 openpyxl.chart.Reference() 函数并传入 3 个参数创建的:
# 1．包含图表数据的 Worksheet 对象
# 2．两个整数的元组，代表矩形选择区域的左上角单元格，该区域包含图表数据：
#   元组中第一个整数是行，第二个整数是列。请注意第一行是 1，不是 0
# 3．两个整数的元组，代表矩形选择区域的右下角单元格，该区域包含图表数据：
#   元组中第一个整数是行，第二个整数是列

wb = openpyxl.Workbook()
sheet = wb.active

for i in range(1, 11):
    sheet['A' + str(i)] = i

refObj = openpyxl.chart.Reference(sheet, min_row=1, min_col=1, max_row=10, max_col=1)
seriesObj = openpyxl.chart.Series(refObj, title='First series')
chartObj = openpyxl.chart.BarChart()
chartObj.append(seriesObj)
chartObj.title = 'My Chart'
# chartObj.drawing.top = 50
# chartObj.drawing.left = 100
# chartObj.drawing.width = 300
# chartObj.drawing.height = 200

sheet.add_chart(chartObj, 'C5')
wb.save('sampleChart.xlsx')
