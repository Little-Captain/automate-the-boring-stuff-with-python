#!/usr/bin/env python

# 使用 openpyxl 模块

import openpyxl
import pprint
import sys

# 打开 excel 文档
wb = openpyxl.load_workbook('example.xlsx')
pprint.pprint(type(wb))

print('----------------------')

# 获取工作表
sheetNames = wb.sheetnames
pprint.pprint(sheetNames)

print('----------------------')

if len(sheetNames) == 0:
    sys.exit()

sheet = wb[sheetNames[0]]
pprint.pprint(type(sheet))
pprint.pprint(sheet.title)

print('----------------------')

# 获取活动表
# 活动表是工作簿在 Excel 中打开时出现的工作表
anotherSheet = wb.active
pprint.pprint(type(anotherSheet))
pprint.pprint(anotherSheet.title)

# 从表格中取得单元格
cell = sheet['A1']  # (A, 1)
pprint.pprint(type(cell))
pprint.pprint(cell.value)
cell = sheet['B1']
pprint.pprint(type(cell))
pprint.pprint(cell.value)

print('----------------------')

# Row 从 1 开始
# Column 从 A 开始
print('Row %s, Column %s is %s' % (str(cell.row), cell.column, cell.value))
print('coordinate %s' % str(cell.coordinate))

print('----------------------')

cell = sheet.cell(row=1, column=1)
pprint.pprint(type(cell))
pprint.pprint(cell.value)

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

# 确定 sheet 大小

print('----------------------')

print('row %d column %d' % (sheet.max_row, sheet.max_column))

print('----------------------')

for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print('({}, {}) {}'.format(i, j, sheet.cell(i, j).value))

print('----------------------')

# 列字母和数字之间的转换
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))
print(get_column_letter(sheet.max_column))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))

print('----------------------')

# 从表中取得行和列
# 可以将 Worksheet 对象切片，取得电子表格中一行、一列或一个矩形区域中的所有 Cell 对象
# 然后可以循环遍历这个切片中的所有单元格
pprint.pprint(tuple(sheet['A1':'C3']))
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

print('----------------------')

# sheet.columns 返回的是生成器对象
# 使用 list 转换为列表对象
columns = list(sheet.columns)
pprint.pprint(columns[1])

for cell in columns[1]:
    print(cell.value)

print(columns[1][1].value)

rows = list(sheet.rows)
pprint.pprint(rows[1])

# 复习
# 工作簿、工作表、单元格
# 1．导入 openpyxl 模块
# 2．调用 openpyxl.load_workbook() 函数
# 3．取得 Workbook 对象
# 4．调用 get_active_sheet() 或 get_sheet_by_name() 工作簿方法
# 5．取得 Worksheet 对象
# 6．使用索引或工作表的 cell() 方法，带上 row 和 column 关键字参数
# 7．取得 Cell 对象
# 8．读取 Cell 对象的 value 属性

print('-----------------------------')

wb.close()

# 写入 Excel 文档

# 创建并保存 Excel 文档
import openpyxl
wb = openpyxl.Workbook()

print(wb.sheetnames)
print(wb.active)
print(wb.active.title)
wb.active.title = 'Spam Bacon Eggs Sheet'
print(wb.active.title)

# 当修改 Workbook 对象或它的工作表和单元格时，电子表格文件不会保存，
# 除非你调用 save() 工作簿方法
# wb.save('test.xlsx')

# 当你编辑从文件中加载的一个电子表格时，总是应该将新的、编辑过的电子表格
# 保存到不同的文件名中。这样，如果代码中有缺陷，导致新的保存到文件中数据
# 不对或讹误，还有最初的电子表格文件可以处理。

print('--------------------')

# 创建和删除工作表
# create_sheet() remove_sheet()
wb = openpyxl.Workbook()
print(wb.sheetnames)
wb.create_sheet()
print(wb.sheetnames)
# create_sheet() 方法返回一个新的 Worksheet 对象，名为 SheetX，
# 它默认是工作簿的最后一个工作表。或者，可以利用 index 和 title 关键字参数，
# 指定新工作表的索引或名称
wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames)
wb.create_sheet(index=2, title='Middle Sheet')
print(wb.sheetnames)
del wb['Middle Sheet']
print(wb.sheetnames)
wb.remove(wb['Sheet1'])
print(wb.sheetnames)
# remove_sheet() 方法接受一个 Worksheet 对象作为其参数，而不是工作表名称的字符串
# 如果你只知道要删除的工作表的名称，就调用 get_sheet_by_name()，将它的返回值传入
# remove_sheet()

# wb.save('test.xlsx')

print('--------------------')

# 将值写入单元格
# 将值写入单元格，很像将值写入字典中的键
# wb = openpyxl.load_workbook('test.xlsx')
sheet = wb.active
# 都能存入数据
sheet['A1'] = 'Hello world!'
sheet['A2'].value = 'Hello'
sheet['A3'] = 'Hello Python'

wb.save('test.xlsx')

# 在原 excel 被加载的情况下，不能保存覆盖。
# 一般需要两个 workbook
# 1. 加载需要处理的 workbook, 只读
# 2. 创建一个新的 workbook, 处理后的数据存入这个 workbook
# 关闭 (1) 中的 workbook，保存 (2) 中的 workbook.
# 如果需要覆盖，使用同一个文件名即可

wb.close()